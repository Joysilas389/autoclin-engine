"""
AutoClin Engine Excel Workbook Generator
Multi-worksheet Excel export with summary, profiles, anomalies,
cleaning actions, method comparison, and before-vs-after metrics.
"""
from typing import Optional
import io


class ExcelGenerator:
    """Generates multi-sheet Excel workbook from pipeline results."""

    def generate(self, result, output_path: str):
        """
        Generate a complete Excel workbook.
        
        Args:
            result: PipelineResult object
            output_path: path to write .xlsx file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise RuntimeError("openpyxl required for Excel generation")

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ──
        ws = wb.active
        ws.title = "Summary"
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="3B82F6")

        summary_data = [
            ("AutoClin Engine Analysis Summary", ""),
            ("", ""),
            ("Dataset Rows", result.total_rows),
            ("Total Anomalies", result.total_anomalies),
            ("Noise Percentage", f"{result.noise_percentage:.2f}%"),
            ("Trust Score", f"{result.trust_score:.1f}/100"),
            ("Selection Mode", result.selection_mode),
            ("Selected Methods", ", ".join(result.selected_methods)),
            ("Duration", f"{result.duration_ms / 1000:.2f}s"),
        ]
        for row_data in summary_data:
            ws.append(row_data)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35

        # ── Sheet 2: Anomaly Log ──
        ws2 = wb.create_sheet("Anomaly Log")
        headers = ["Row", "Type", "Severity", "Confidence", "Flagged Columns", "Rationale"]
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        sev_colors = {
            "critical": PatternFill("solid", fgColor="FEE2E2"),
            "high": PatternFill("solid", fgColor="FEF3C7"),
            "medium": PatternFill("solid", fgColor="FEF9C3"),
            "low": PatternFill("solid", fgColor="DCFCE7"),
        }
        for cls in result.anomaly_classifications:
            row = [
                cls.row_index,
                cls.anomaly_type.replace("_", " "),
                cls.severity,
                round(cls.confidence, 3),
                "; ".join(cls.flagged_columns[:3]),
                cls.rationale[:100],
            ]
            ws2.append(row)
            fill = sev_colors.get(cls.severity)
            if fill:
                for cell in ws2[ws2.max_row]:
                    cell.fill = fill

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws2.column_dimensions[col_letter].width = 20

        # ── Sheet 3: Method Comparison ──
        ws3 = wb.create_sheet("Method Comparison")
        ws3.append(["Method", "Composite", "NDC", "AD", "SS", "CP", "EX", "CC", "Anomalies", "Duration(ms)", "Selected"])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
        for mr in result.method_rankings:
            ws3.append([
                mr["method"], mr["composite"], mr["ndc"], mr["ad"], mr["ss"],
                mr["cp"], mr["ex"], mr["cc"], mr.get("anomaly_count", 0),
                mr.get("duration_ms", 0), "Yes" if mr.get("selected") else "",
            ])

        # ── Sheet 4: Cleaning Recommendations ──
        ws4 = wb.create_sheet("Cleaning Actions")
        ws4.append(["Row", "Column", "Type", "Severity", "Action", "Risk", "Rationale"])
        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill
        for rec in result.cleaning_recommendations:
            ws4.append([
                rec.row_index, rec.column_name, rec.anomaly_type.replace("_", " "),
                rec.severity, rec.recommended_action, rec.auto_clean_risk,
                rec.rationale[:100],
            ])

        # ── Sheet 5: Quality Metrics ──
        ws5 = wb.create_sheet("Quality Metrics")
        ws5.append(["Metric", "Before", "After" if result.after_metrics else ""])
        for cell in ws5[1]:
            cell.font = header_font
            cell.fill = header_fill
        bm = result.before_metrics
        am = result.after_metrics or {}
        metrics = [
            ("Noise %", bm.get("noise_pct"), am.get("noise_pct", "")),
            ("Trust Score", bm.get("trust_score"), am.get("trust_score", "")),
            ("Missingness %", bm.get("missingness_pct"), am.get("missingness_pct", "")),
            ("Duplicate %", bm.get("duplicate_pct"), am.get("duplicate_pct", "")),
            ("Plausibility %", bm.get("plausibility_rate"), am.get("plausibility_rate", "")),
        ]
        for m in metrics:
            ws5.append(m)

        # ── Sheet 6: Audit Trail ──
        if result.audit_entries:
            ws6 = wb.create_sheet("Audit Trail")
            ws6.append(["TXN ID", "Timestamp", "Row", "Column", "Original", "New", "Action", "Risk"])
            for cell in ws6[1]:
                cell.font = header_font
                cell.fill = header_fill
            for entry in result.audit_entries:
                ws6.append([
                    entry.get("transaction_id"), entry.get("timestamp"),
                    entry.get("row_index"), entry.get("column_name"),
                    entry.get("original_value"), entry.get("new_value"),
                    entry.get("action"), entry.get("risk_level"),
                ])

        wb.save(output_path)
        return output_path
